"""Gera a base de dados para o dashboard de capacitacao (ENAP).

Dimensoes do dashboard:
  1. categoria -> IA / Dados / Gestao / Outros  (coluna `categoria` de
                cursos_alvo.csv; join por nome_curso normalizado). O painel de
                capacitacao conta IA+Dados; Gestao e Outros ficam na base
                (auditaveis) mas fora dos indicadores.
  2. esfera  -> Federal / Estadual / Municipal / "(sem esfera)"
  3. setor   -> "Publico"  quando esfera in {Federal, Estadual, Municipal}
                "Privado/Nao-servidor"  quando esfera vazia
  4. poder   -> Executivo / Legislativo / Judiciario / "(sem poder)"
                (so preenchido p/ parte dos servidores publicos)
  5. tempo   -> ano_mes (YYYY-MM), ano (int), mes (int)  [de dt_matricula]

  Nota: a fonte da ENAP NAO tem campo de "iniciativa privada". O unico sinal
  de vinculo publico e a coluna `esfera` (so preenchida para servidores
  publicos). Quem nao e servidor (empresa privada, sociedade civil, publico
  geral) fica com esfera vazia. O setor "Privado/Nao-servidor" agrega todos
  esses casos -- nao e "empresa privada" pura. `poder` so vem preenchido para
  ~68% dos registros publicos; o restante fica "(sem poder)".

Metricas (calculadas a partir da tabela fato):
  - matriculas      = contagem de linhas
  - pessoas_unicas  = codigo_pessoa distintos (NAO somam entre fatias -- por
                      isso a tabela fato e necessaria; a agregada e atalho)

Filtros: sit_matricula='Concluida'; apenas os 35 cursos alvo;
janela [INICIO_HISTORICO, --ate] sobre dt_matricula.

Saidas (em docs/):
  - dashboard_base.csv          tabela fato: 1 linha por matricula concluida
  - dashboard_agregado.csv      agregado por (tempo x categoria x esfera x setor x poder)
  - relatorio_capacitacao_ia.html  relatorio estatico (4 indicadores x grupos)
  - relatorio_capacitacao_ia.xlsx  mesma tabela do relatorio, formatada (Excel)

Reusa os helpers de coleta/normalizacao de atualizar_historico.py para que a
juncao por nome de curso seja identica a do pipeline mensal.
"""

from __future__ import annotations

import argparse
import csv
import io
import shutil
import sys
from datetime import date, datetime
from html import escape
from pathlib import Path

import pandas as pd

# Garante que atualizar_historico.py (mesmo diretorio) seja importavel
# independentemente do diretorio de invocacao.
sys.path.insert(0, str(Path(__file__).resolve().parent))

from atualizar_historico import (  # noqa: E402
    INICIO_HISTORICO,
    URL_DEFAULT,
    extrair_ano_mes,
    iter_csvs,
    normalize,
    ultimo_mes_completo,
)

REPO_ROOT = Path(__file__).resolve().parents[1]
DOCS_DIR = REPO_ROOT / "docs"
CURSOS_ALVO = Path(__file__).parent / "cursos_alvo.csv"
BASE_CSV = DOCS_DIR / "dashboard_base.csv"
BACKUP_DIR = DOCS_DIR / ".base_backups"
AGREGADO_CSV = DOCS_DIR / "dashboard_agregado.csv"
RELATORIO_HTML = DOCS_DIR / "relatorio_capacitacao_ia.html"
RELATORIO_XLSX = DOCS_DIR / "relatorio_capacitacao_ia.xlsx"

ESFERAS_PUBLICAS = {"Federal", "Estadual", "Municipal"}
SEM_ESFERA = "(sem esfera)"
PODERES = {"Executivo", "Legislativo", "Judiciario"}
SEM_PODER = "(sem poder)"


CATEGORIAS_VALIDAS = {"IA", "Dados", "Gestão", "Outros"}
# Categorias que compoem o painel de capacitacao em IA e Dados (indicadores).
CATEGORIAS_PAINEL = {"IA", "Dados"}


def carregar_alvos() -> dict[str, tuple[int, str, str]]:
    """Le cursos_alvo.csv -> {nome_norm: (id_curso, nome, categoria)}.

    Exige a coluna `categoria` (IA / Dados / Gestão / Outros) criada para a
    classificacao dos cursos-alvo.
    """
    df = pd.read_csv(CURSOS_ALVO)
    faltando = {"id_curso", "tx_nome_curso", "categoria"} - set(df.columns)
    if faltando:
        raise SystemExit(
            f"cursos_alvo.csv sem coluna(s) {sorted(faltando)} -- "
            "rode a classificacao por categoria antes de gerar a base."
        )
    invalidas = set(df["categoria"].dropna().unique()) - CATEGORIAS_VALIDAS
    if invalidas:
        raise SystemExit(
            f"cursos_alvo.csv com categoria(s) invalida(s): {sorted(invalidas)} -- "
            f"use apenas {sorted(CATEGORIAS_VALIDAS)}."
        )
    df["nome_norm"] = df["tx_nome_curso"].map(normalize)
    return {
        row.nome_norm: (int(row.id_curso), row.tx_nome_curso, str(row.categoria))
        for row in df.itertuples(index=False)
    }


def classificar(esfera_raw: str | None, poder_raw: str | None) -> tuple[str, str, str]:
    """(esfera_raw, poder_raw) -> (esfera_dim, setor, poder_dim).

    poder so e considerado quando o registro e publico (esfera preenchida);
    para privado/nao-servidor o poder fica "(sem poder)".
    """
    e = (esfera_raw or "").strip()
    if e in ESFERAS_PUBLICAS:
        p = (poder_raw or "").strip()
        return e, "Publico", (p if p in PODERES else SEM_PODER)
    return SEM_ESFERA, "Privado/Nao-servidor", SEM_PODER


def _set_csv_field_size_limit() -> None:
    """Maior field_size_limit suportado pela plataforma.

    No Python do Windows o C long e 32-bit, entao sys.maxsize (2^63-1)
    estoura em csv.field_size_limit. Reduz ate caber.
    """
    limite = sys.maxsize
    while True:
        try:
            csv.field_size_limit(limite)
            return
        except OverflowError:
            limite //= 10


def processar_csv(
    stream: io.TextIOBase, mapa: dict[str, tuple[int, str, str]]
) -> list[dict]:
    """Extrai linhas-fato (Concluida + 35 cursos alvo) de um CSV mensal."""
    _set_csv_field_size_limit()
    reader = csv.DictReader(stream, delimiter="|")
    linhas: list[dict] = []
    for row in reader:
        if (row.get("sit_matricula") or "").strip() != "Concluida":
            continue
        nn = normalize(row.get("nome_curso") or "")
        alvo = mapa.get(nn)
        if alvo is None:
            continue
        dt = (row.get("dt_matricula") or "")[:7]
        if len(dt) != 7 or dt[4] != "-":
            continue
        esfera, setor, poder = classificar(row.get("esfera"), row.get("poder"))
        id_curso, nome, categoria = alvo
        linhas.append(
            {
                "ano_mes": dt,
                "ano": int(dt[:4]),
                "mes": int(dt[5:7]),
                "id_curso": id_curso,
                "nome_curso": nome,
                "categoria": categoria,
                "esfera": esfera,
                "setor": setor,
                "poder": poder,
                "codigo_pessoa": (row.get("codigo_pessoa") or "").strip(),
            }
        )
    return linhas


def coletar(
    source: str | Path, janela_max: str, mapa: dict[str, tuple[int, str, str]]
) -> pd.DataFrame:
    todas: list[dict] = []
    for nome_logico, stream in iter_csvs(source):
        ym = extrair_ano_mes(nome_logico)
        if ym is not None and (ym < INICIO_HISTORICO or ym > janela_max):
            print(f"Pulando {nome_logico} (fora da janela)", flush=True)
            stream.close()
            continue
        print(f"Processando {nome_logico}", flush=True)
        todas.extend(processar_csv(stream, mapa))
        stream.close()

    cols = [
        "ano_mes", "ano", "mes", "id_curso", "nome_curso",
        "categoria", "esfera", "setor", "poder", "codigo_pessoa",
    ]
    df = pd.DataFrame(todas, columns=cols)
    if df.empty:
        return df
    df = df[(df["ano_mes"] >= INICIO_HISTORICO) & (df["ano_mes"] <= janela_max)]
    return df.sort_values(["ano_mes", "id_curso"]).reset_index(drop=True)


FATO_COLS = [
    "ano_mes", "ano", "mes", "id_curso", "nome_curso",
    "categoria", "esfera", "setor", "poder", "codigo_pessoa",
]
FATO_DTYPES = {
    "ano_mes": str, "ano": "Int64", "mes": "Int64", "id_curso": "Int64",
    "nome_curso": str, "categoria": str, "esfera": str, "setor": str,
    "poder": str, "codigo_pessoa": str,
}


def merge_fato(novo: pd.DataFrame, base_csv: Path) -> pd.DataFrame:
    """Acumula a tabela-fato preservando meses fora da janela do download.

    Idempotente por mês (mesma lógica de merge_historico em
    atualizar_historico.py): para cada ano_mes presente em `novo`, substitui
    as linhas correspondentes na base existente; meses antigos que não vieram
    no download atual são preservados. Isso permite que o dashboard acumule
    histórico mesmo com a fonte ENAP rolante (últimos 12 meses).
    """
    if not base_csv.exists():
        return novo
    antigo = pd.read_csv(base_csv, dtype=FATO_DTYPES)[FATO_COLS]
    antigo["codigo_pessoa"] = antigo["codigo_pessoa"].fillna("")
    if novo.empty:
        return antigo.sort_values(["ano_mes", "id_curso"]).reset_index(drop=True)
    meses_novos = set(novo["ano_mes"].unique())
    antigo_filtrado = antigo[~antigo["ano_mes"].isin(meses_novos)]
    final = pd.concat([antigo_filtrado, novo], ignore_index=True)
    # normaliza tipos (concat com Int64 vs int)
    for c in ("ano", "mes", "id_curso"):
        final[c] = final[c].astype("int64")
    final["categoria"] = final["categoria"].astype(str)
    final["codigo_pessoa"] = final["codigo_pessoa"].fillna("").astype(str)
    return final.sort_values(["ano_mes", "id_curso"]).reset_index(drop=True)


def _meses_da_base(base_csv: Path) -> set[str]:
    """Conjunto de ano_mes presentes numa base existente (vazio se nao existe)."""
    if not base_csv.exists():
        return set()
    try:
        antiga = pd.read_csv(base_csv, usecols=["ano_mes"], dtype={"ano_mes": str})
    except (ValueError, pd.errors.EmptyDataError):
        return set()
    return set(antiga["ano_mes"].dropna().unique())


def _backup_base(base_csv: Path) -> Path:
    """Copia base_csv para docs/.base_backups/<nome>-<carimbo>.csv antes de sobrescrever.

    Backup datado (nao sobrescreve backups anteriores) para que um run ruim
    -- ex.: download incompleto -- nunca destrua a ultima base boa.
    """
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    carimbo = datetime.now().strftime("%Y%m%d-%H%M%S")
    destino = BACKUP_DIR / f"{base_csv.stem}-{carimbo}.csv"
    shutil.copy2(base_csv, destino)
    return destino


AGG_DIMS = ["ano_mes", "ano", "mes", "categoria", "esfera", "setor", "poder"]


def agregar(fato: pd.DataFrame) -> pd.DataFrame:
    """Agrega por (tempo, categoria, esfera, setor, poder): matriculas + pessoas_unicas.

    pessoas_unicas usa contagem distinta por celula -- NAO e somavel entre
    celulas (uma pessoa pode aparecer em mais de uma combinacao, inclusive em
    meses diferentes). Para o total do periodo, conte distintos na tabela fato.
    """
    if fato.empty:
        return pd.DataFrame(columns=AGG_DIMS + ["matriculas", "pessoas_unicas"])
    cp = fato["codigo_pessoa"].replace("", pd.NA)
    g = fato.assign(codigo_pessoa=cp).groupby(AGG_DIMS, dropna=False)
    out = g.agg(
        matriculas=("codigo_pessoa", "size"),
        pessoas_unicas=("codigo_pessoa", "nunique"),
    ).reset_index()
    return out.sort_values(
        ["ano_mes", "categoria", "setor", "esfera", "poder"]
    ).reset_index(drop=True)


def _fmt(n: int) -> str:
    """Inteiro com separador de milhar pt-BR (ponto)."""
    return f"{int(n):,}".replace(",", ".")


def _indicadores(sub: pd.DataFrame) -> dict[str, int]:
    """Os indicadores do relatorio para um subconjunto da tabela fato.

    O painel de capacitacao cobre IA+Dados; Gestao/Outros ficam de fora dos
    indicadores (mas seguem na base). Cada metrica e quebrada em IA e Dados.
    """
    com_pessoa = sub["codigo_pessoa"] != ""
    ia = sub["categoria"] == "IA"
    dados = sub["categoria"] == "Dados"
    painel = ia | dados
    return {
        "pessoas": sub.loc[com_pessoa & painel, "codigo_pessoa"].nunique(),
        "pessoas_ia": sub.loc[com_pessoa & ia, "codigo_pessoa"].nunique(),
        "pessoas_dados": sub.loc[com_pessoa & dados, "codigo_pessoa"].nunique(),
        "matriculas": int(painel.sum()),
        "matriculas_ia": int(ia.sum()),
        "matriculas_dados": int(dados.sum()),
    }


def _linhas_relatorio(fato: pd.DataFrame) -> list[tuple[str, str, dict | None]]:
    pub = fato[fato["setor"] == "Publico"]
    priv = fato[fato["setor"] == "Privado/Nao-servidor"]
    espec = [
        ("secao", "Total no Setor Público", pub),
        ("subhdr", "Por Esfera no Setor Público", None),
        ("sub", "Federal", fato[fato["esfera"] == "Federal"]),
        ("sub", "Estadual", fato[fato["esfera"] == "Estadual"]),
        ("sub", "Municipal", fato[fato["esfera"] == "Municipal"]),
        ("subhdr", "Por Poder no Setor Público", None),
        ("sub", "Executivo", pub[pub["poder"] == "Executivo"]),
        ("sub", "Legislativo", pub[pub["poder"] == "Legislativo"]),
        ("sub", "Judiciário", pub[pub["poder"] == "Judiciario"]),
        ("sub muted", "(Poder não informado)", pub[pub["poder"] == SEM_PODER]),
        ("secao", "No Setor Privado / não-servidor", priv),
        ("total", "Total (independente de esfera e poder)", fato),
    ]
    return [
        (cls, rot, None if sub is None else _indicadores(sub))
        for cls, rot, sub in espec
    ]


def gerar_relatorio_html(fato: pd.DataFrame, janela_max: str) -> str:
    data_hoje = date.today().strftime("%d/%m/%Y")
    meses = sorted(fato["ano_mes"].unique())
    periodo = f"{meses[0]} a {meses[-1]}" if meses else "—"
    cursos_por_cat = fato.groupby("categoria")["id_curso"].nunique()
    n_cursos_ia = int(cursos_por_cat.get("IA", 0))
    n_cursos_dados = int(cursos_por_cat.get("Dados", 0))
    n_cursos = fato["id_curso"].nunique()
    matr_painel = int(fato["categoria"].isin(CATEGORIAS_PAINEL).sum())

    # Linhas da tabela principal (indicadores IA/Dados x grupos)
    corpo = []
    for cls, rotulo, ind in _linhas_relatorio(fato):
        if ind is None:  # cabecalho de subsecao
            corpo.append(
                f'<tr class="subhdr"><td colspan="5">{escape(rotulo)}</td></tr>'
            )
            continue
        corpo.append(
            f'<tr class="{cls}"><td>{escape(rotulo)}</td>'
            f'<td>{_fmt(ind["matriculas_ia"])}</td>'
            f'<td>{_fmt(ind["matriculas_dados"])}</td>'
            f'<td>{_fmt(ind["pessoas_ia"])}</td>'
            f'<td>{_fmt(ind["pessoas_dados"])}</td></tr>'
        )
    tabela_ind = "".join(corpo)

    # Tabela auxiliar: detalhe por curso (id, nome, categoria, matriculas, pessoas)
    cp = fato["codigo_pessoa"].replace("", pd.NA)
    por_curso = (
        fato.assign(_cp=cp)
        .groupby(["id_curso", "nome_curso", "categoria"], as_index=False)
        .agg(matriculas=("_cp", "size"), pessoas=("_cp", "nunique"))
        .sort_values("id_curso")
    )
    linhas_curso = "".join(
        f'<tr><td>{r.id_curso}</td><td class="esq">{escape(r.nome_curso)}</td>'
        f'<td class="esq">{escape(r.categoria)}</td>'
        f'<td>{_fmt(r.matriculas)}</td><td>{_fmt(r.pessoas)}</td></tr>'
        for r in por_curso.itertuples(index=False)
    )

    return f"""<!doctype html>
<html lang="pt-br">
<head>
<meta charset="utf-8">
<title>Relatório — Capacitação em IA (ENAP)</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; max-width: 1100px; margin: 2rem auto; padding: 0 1rem; color: #222; }}
  h1 {{ margin-bottom: 0.2rem; }}
  h2 {{ margin-top: 2rem; }}
  .meta {{ color: #666; margin-bottom: 1.5rem; font-size: 0.9rem; }}
  table {{ border-collapse: collapse; margin: 1rem 0; font-size: 0.9rem; width: 100%; }}
  th, td {{ border: 1px solid #ddd; padding: 6px 10px; text-align: right; }}
  th {{ background: #f4f4f4; }}
  td:first-child, th:first-child, .esq {{ text-align: left; }}
  tr.secao td {{ font-weight: 700; background: #eef4fb; }}
  tr.total td {{ font-weight: 700; background: #e9f7ef; border-top: 2px solid #999; }}
  tr.subhdr td {{ font-weight: 600; background: #fafafa; color: #555; text-align: left; }}
  tr.sub td:first-child {{ padding-left: 1.8rem; }}
  tr.muted td {{ color: #999; }}
  .scroll {{ overflow-x: auto; }}
  a {{ color: #0366d6; }}
  .legenda {{ font-size: 0.85rem; color: #555; }}
</style>
</head>
<body>
<h1>Relatório sobre Conclusão de Cursos dos Programas de Capacitação em IA e Dados</h1>
<p class="meta">
  <strong>Data:</strong> {data_hoje} ·
  <strong>Período coberto:</strong> {periodo} (matrículas concluídas) ·
  Fonte: <a href="https://dadosaberto.evg.gov.br/">dadosaberto.evg.gov.br</a><br>
  Programa = {n_cursos} cursos-alvo · Painel de capacitação = {n_cursos_ia} cursos de IA +
  {n_cursos_dados} de Dados ({_fmt(matr_painel)} matrículas). Cursos de Gestão e Outros
  seguem na base, mas ficam fora dos indicadores.
</p>

<h2>Indicadores por grupo</h2>
<div class="scroll">
<table>
<thead><tr>
  <th>Grupo</th>
  <th>Matrículas concluídas<br>(cursos de IA)</th>
  <th>Matrículas concluídas<br>(cursos de Dados)</th>
  <th>Pessoas concluintes<br>(≥1 curso de IA)</th>
  <th>Pessoas concluintes<br>(≥1 curso de Dados)</th>
</tr></thead>
<tbody>{tabela_ind}</tbody>
</table>
</div>
<p class="legenda">
  <strong>Indicadores cobrem apenas IA e Dados</strong> (cursos de Gestão e Outros
  ficam de fora). <strong>Pessoas</strong> = servidores/pessoas distintas (contagem
  única de <code>codigo_pessoa</code>); <strong>não somam entre grupos</strong> — uma
  pessoa pode aparecer em mais de um recorte e em mais de um curso.
  <strong>Matrículas</strong> = nº de conclusões (somam).
  Federal+Estadual+Municipal reconciliam o total público em matrículas;
  os poderes não, pois ~32% dos registros públicos vêm sem poder informado
  (linha "Poder não informado").
  "Setor Privado / não-servidor" agrega iniciativa privada, sociedade civil e
  público geral — a fonte não distingue empresa privada pura.
</p>

<h2>Detalhe por curso</h2>
<div class="scroll">
<table>
<thead><tr>
  <th>id</th><th class="esq">Curso</th><th class="esq">Categoria</th>
  <th>Matrículas concl.</th><th>Pessoas concl.</th>
</tr></thead>
<tbody>{linhas_curso}</tbody>
</table>
</div>

<p class="meta">Filtros: <code>sit_matricula = 'Concluida'</code> ·
{n_cursos} cursos-alvo · <code>dt_matricula</code> em [{INICIO_HISTORICO}, {janela_max}] ·
junção por <code>nome_curso</code> normalizado. Base de dados reaproveitável:
<a href="dashboard_base.csv">dashboard_base.csv</a> (tabela fato) ·
<a href="dashboard_agregado.csv">dashboard_agregado.csv</a>.</p>
</body>
</html>
"""


COLS_RELATORIO = [
    "Grupo",
    "Matrículas concluídas (cursos de IA)",
    "Matrículas concluídas (cursos de Dados)",
    "Pessoas concluintes (≥1 curso de IA)",
    "Pessoas concluintes (≥1 curso de Dados)",
]


def gerar_relatorio_xlsx(fato: pd.DataFrame, janela_max: str) -> None:
    """Escreve RELATORIO_XLSX com a tabela de 4 indicadores x grupos (formatada).

    openpyxl e importado aqui para nao virar dependencia obrigatoria de quem
    so quer os CSVs/HTML.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    meses = sorted(fato["ano_mes"].unique())
    periodo = f"{meses[0]} a {meses[-1]}" if meses else "—"
    data_hoje = date.today().strftime("%d/%m/%Y")

    fill_hdr = PatternFill("solid", fgColor="DDDDDD")
    fill_secao = PatternFill("solid", fgColor="EEF4FB")
    fill_sub = PatternFill("solid", fgColor="FAFAFA")
    fill_total = PatternFill("solid", fgColor="E9F7EF")
    thin = Side(style="thin", color="CCCCCC")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Indicadores"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Relatório sobre Conclusão de Cursos dos Programas de Capacitação em IA e Dados"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Data: {data_hoje}  ·  Período coberto: {periodo}  ·  Fonte: dadosaberto.evg.gov.br"
    ws["A2"].font = Font(size=10, color="666666")

    hdr_row = 4
    for col, titulo in enumerate(COLS_RELATORIO, start=1):
        c = ws.cell(row=hdr_row, column=col, value=titulo)
        c.font = Font(bold=True)
        c.fill = fill_hdr
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="left" if col == 1 else "right")
        c.border = borda

    r = hdr_row + 1
    for cls, rotulo, ind in _linhas_relatorio(fato):
        if ind is None:  # cabecalho de subsecao
            ws.cell(row=r, column=1, value=rotulo)
            for col in range(1, 6):
                cell = ws.cell(row=r, column=col)
                cell.fill = fill_sub
                cell.border = borda
                if col == 1:
                    cell.font = Font(bold=True, italic=True, color="555555")
            r += 1
            continue

        valores = [rotulo, ind["matriculas_ia"], ind["matriculas_dados"],
                   ind["pessoas_ia"], ind["pessoas_dados"]]
        for col, v in enumerate(valores, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.border = borda
            if col == 1:
                cell.alignment = Alignment(
                    horizontal="left", indent=2 if cls.startswith("sub") else 0
                )
            else:
                cell.number_format = "#,##0"
            if cls == "secao":
                cell.font = Font(bold=True)
                cell.fill = fill_secao
            elif cls == "total":
                cell.font = Font(bold=True)
                cell.fill = fill_total
            elif "muted" in cls:
                cell.font = Font(color="999999")
        r += 1

    widths = [44, 22, 22, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[hdr_row].height = 46
    ws.freeze_panes = f"A{hdr_row + 1}"

    nota = (
        'Indicadores cobrem apenas IA e Dados (Gestão e Outros ficam de fora). '
        'Pessoas = contagem única de pessoas (codigo_pessoa); NÃO somam entre '
        'grupos. Matrículas = nº de conclusões (somam). Federal+Estadual+'
        'Municipal reconciliam o total público em matrículas; os poderes não '
        '(~32% dos registros públicos sem poder informado). "Setor Privado / '
        'não-servidor" agrega iniciativa privada, sociedade civil e público '
        f'geral. Filtros: sit_matricula=Concluida; {fato["id_curso"].nunique()} '
        f'cursos-alvo; dt_matricula em [{INICIO_HISTORICO}, {janela_max}].'
    )
    ws.cell(row=r + 1, column=1, value=nota).font = Font(size=9, color="777777")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)
    ws.cell(row=r + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r + 1].height = 70

    wb.save(RELATORIO_XLSX)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source",
        default=URL_DEFAULT,
        help="URL ou caminho local (tar.gz, CSV ou diretorio). Default: dump ENAP.",
    )
    parser.add_argument(
        "--ate",
        default=None,
        help="Ultimo ano-mes a processar (YYYY-MM). Default: mes anterior ao atual.",
    )
    parser.add_argument(
        "--rebuild",
        action="store_true",
        help="Reconstroi a base do zero (NAO acumula). Default: acumula meses "
             "antigos preservando dashboard_base.csv (fonte ENAP e rolante).",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Confirma a sobrescrita mesmo que a nova base perca meses que "
             "existem na base atual (a salvaguarda aborta por padrao).",
    )
    args = parser.parse_args()

    janela_max = args.ate or ultimo_mes_completo()
    print(f"Janela: {INICIO_HISTORICO} ate {janela_max}")

    mapa = carregar_alvos()
    from collections import Counter
    cats = Counter(v[2] for v in mapa.values())
    print(f"Cursos alvo: {len(mapa)} ("
          + ", ".join(f"{k}={cats.get(k, 0)}" for k in
                      ("IA", "Dados", "Gestão", "Outros"))
          + ")")

    fato = coletar(args.source, janela_max, mapa)
    if not args.rebuild:
        antes = len(fato)
        fato = merge_fato(fato, BASE_CSV)
        meses = sorted(fato["ano_mes"].unique()) if not fato.empty else []
        print(f"Acumulado: {antes} linhas novas -> {len(fato)} linhas totais "
              f"({len(meses)} meses: {meses[0] if meses else '-'}..{meses[-1] if meses else '-'})")
    agg = agregar(fato)

    # Salvaguarda anti-encolhimento: a fonte normal da ENAP e rolante (ultimos
    # 12 meses). Rodar sem o dump historico completo -- sobretudo com --rebuild
    # -- apagaria meses antigos da base. Aborta se a nova base perderia meses
    # presentes na atual (a menos que --force confirme) e sempre faz backup
    # datado antes de sobrescrever.
    meses_antigos = _meses_da_base(BASE_CSV)
    meses_novos = set(fato["ano_mes"].unique()) if not fato.empty else set()
    perdidos = sorted(meses_antigos - meses_novos)
    if perdidos and not args.force:
        print(
            f"\nABORTADO: a nova base perderia {len(perdidos)} mes(es) que existem "
            f"na base atual ({perdidos[0]}..{perdidos[-1]}).\n"
            f"  Base atual: {len(meses_antigos)} meses | nova: {len(meses_novos)} meses.\n"
            "  Isso costuma significar que a fonte e a janela rolante (ultimos 12\n"
            "  meses), nao o dump historico completo.\n"
            "    - Para acumular preservando o historico, rode SEM --rebuild.\n"
            "    - Se a perda for intencional, repita com --force.",
            file=sys.stderr,
        )
        return 1

    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    if BASE_CSV.exists():
        bkp = _backup_base(BASE_CSV)
        print(f"Backup da base anterior: {bkp.relative_to(REPO_ROOT)}")
    fato.to_csv(BASE_CSV, index=False)
    agg.to_csv(AGREGADO_CSV, index=False)

    if fato.empty:
        print("\nNenhuma linha gerada (verifique --source / janela).")
        RELATORIO_HTML.write_text(gerar_relatorio_html(fato, janela_max), encoding="utf-8")
        return 0

    RELATORIO_HTML.write_text(gerar_relatorio_html(fato, janela_max), encoding="utf-8")
    gerar_relatorio_xlsx(fato, janela_max)

    total_matr = len(fato)
    total_pessoas = fato.loc[fato["codigo_pessoa"] != "", "codigo_pessoa"].nunique()
    cursos_presentes = fato["id_curso"].nunique()
    print(f"\nTabela fato: {total_matr} linhas (matriculas concluidas)")
    print(f"Meses: {sorted(fato['ano_mes'].unique())}")
    print(f"Cursos com conclusao no periodo: {cursos_presentes}/{len(mapa)}")
    print(f"Pessoas unicas (global, distintas): {total_pessoas}")
    print("\nRelatorio (indicadores IA/Dados x grupos):")
    for cls, rotulo, ind in _linhas_relatorio(fato):
        if ind is None:
            print(f"  [{rotulo}]")
            continue
        print(
            f"  {rotulo:42s} matr_IA={ind['matriculas_ia']:>7} "
            f"matr_Dados={ind['matriculas_dados']:>7} "
            f"pess_IA={ind['pessoas_ia']:>7} pess_Dados={ind['pessoas_dados']:>7}"
        )
    print(
        f"\nArquivos: {BASE_CSV}, {AGREGADO_CSV}, {RELATORIO_HTML}, {RELATORIO_XLSX}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
